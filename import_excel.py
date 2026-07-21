import openpyxl
import sys
sys.path.insert(0, '/mnt/d/Agncy/agence_transfert')
from app import app
from models import db, Agency, OperationType, User, VirtualStock, CashBalance, ClotureJournaliere, Excedent
from datetime import datetime

EXCEL_PATH = '/mnt/d/Agncy/CLOTURE SHOP AIRTEL - Copie (Enregistr\u00e9 automatiquement).xlsx'

def safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(',', ''))
        except (ValueError, TypeError):
            return 0.0
    return 0.0

with app.app_context():
    agency = Agency.query.filter_by(name='Agence Principale').first()
    if not agency:
        agency = Agency(name='Agence Principale', address='Bukavu', is_active=True)
        db.session.add(agency)
        db.session.flush()

    users = {}
    for name in ['Riziki', 'Lumiere']:
        u = User.query.filter_by(username=name.lower(), agency_id=agency.id).first()
        if not u:
            u = User(username=name.lower(), full_name=name, role='guichetier', agency_id=agency.id, is_active=True)
            u.set_password('guichetier123')
            db.session.add(u)
            db.session.flush()
        users[name] = u

    db.session.commit()
    print(f'Agency: {agency.name} (id={agency.id})')
    print(f'Users: {[(n, u.id) for n, u in users.items()]}')

    # Clean old closings for this agency
    ClotureJournaliere.query.filter_by(agency_id=agency.id).delete()
    db.session.commit()

    wb = openpyxl.load_workbook(EXCEL_PATH)

    GUICHETIER_MAP = {
        'RIZIKI': users['Riziki'],
        'LUMIERE': users['Lumiere'],
    }

    for sname in wb.sheetnames:
        ws = wb[sname]
        name_key = sname.upper()
        guichetier = GUICHETIER_MAP.get(name_key)
        if not guichetier:
            print(f'Skipping unknown sheet: {sname}')
            continue

        rows_imported = 0
        last_data = {}

        for r in range(4, ws.max_row + 1):
            date_val = ws.cell(r, 1).value
            if not date_val:
                continue
            if isinstance(date_val, datetime):
                dt = date_val.date()
            else:
                continue

            solde_u = safe_float(ws.cell(r, 2).value)
            solde_c = safe_float(ws.cell(r, 3).value)
            ajout_u = safe_float(ws.cell(r, 4).value)
            ajout_c = safe_float(ws.cell(r, 5).value)
            retrait_u = safe_float(ws.cell(r, 6).value)
            retrait_c = safe_float(ws.cell(r, 7).value)
            ex_swaps = safe_float(ws.cell(r, 8).value)
            ex_bureau = safe_float(ws.cell(r, 9).value)
            ex_autres = safe_float(ws.cell(r, 10).value)
            total_exc = ex_swaps + ex_bureau + ex_autres
            creance_u = safe_float(ws.cell(r, 11).value)
            creance_c = safe_float(ws.cell(r, 12).value)
            airtel_u = safe_float(ws.cell(r, 15).value)
            airtel_c = safe_float(ws.cell(r, 16).value)
            vodacom_u = safe_float(ws.cell(r, 17).value)
            vodacom_c = safe_float(ws.cell(r, 18).value)
            orange_u = safe_float(ws.cell(r, 19).value)
            orange_c = safe_float(ws.cell(r, 20).value)
            pepele_u = safe_float(ws.cell(r, 21).value)
            pepele_c = safe_float(ws.cell(r, 22).value)
            cash_u = safe_float(ws.cell(r, 23).value)
            cash_c = safe_float(ws.cell(r, 24).value)
            dettes_u = safe_float(ws.cell(r, 25).value)
            dettes_c = safe_float(ws.cell(r, 26).value)
            taux = safe_float(ws.cell(r, 31).value)

            virtuel_u = airtel_u + vodacom_u + orange_u + pepele_u
            virtuel_c = airtel_c + vodacom_c + orange_c + pepele_c
            total_sin_usd = solde_u + ajout_u - retrait_u
            total_sin_cdf = solde_c + ajout_c - retrait_c + total_exc
            total_actif_usd = virtuel_u + cash_u
            total_actif_cdf = virtuel_c + cash_c
            cumule_u = total_actif_usd - total_sin_usd
            cumule_c = total_actif_cdf - total_sin_cdf
            manque = cumule_c + (cumule_u * taux)

            cloture = ClotureJournaliere(
                agency_id=agency.id,
                guichetier_id=guichetier.id,
                date=dt,
                status='valide',
                taux_change=taux,
                solde_initial_usd=solde_u,
                solde_initial_cdf=solde_c,
                ajout_initial_usd=ajout_u,
                ajout_initial_cdf=ajout_c,
                retrait_initial_usd=retrait_u,
                retrait_initial_cdf=retrait_c,
                total_excedent_cdf=total_exc,
                total_creance_usd=creance_u,
                total_creance_cdf=creance_c,
                total_virtuel_usd=virtuel_u,
                total_virtuel_cdf=virtuel_c,
                total_cash_usd=cash_u,
                total_cash_cdf=cash_c,
                total_dettes_usd=dettes_u,
                total_dettes_cdf=dettes_c,
                total_actif_usd=total_actif_usd,
                total_actif_cdf=total_actif_cdf,
                cumule_usd=cumule_u,
                cumule_cdf=cumule_c,
                manque=manque,
                created_by=1,
            )
            db.session.add(cloture)
            db.session.flush()

            for cat, val in [('SWAPS', ex_swaps), ('Bureau de change', ex_bureau), ('Autres', ex_autres)]:
                if val > 0:
                    db.session.add(Excedent(cloture_id=cloture.id, categorie=cat, montant_cdf=val))

            if airtel_u or airtel_c or vodacom_u or vodacom_c or orange_u or orange_c or pepele_u or pepele_c or cash_u or cash_c:
                last_data = {
                    'airtel_u': airtel_u, 'airtel_c': airtel_c,
                    'vodacom_u': vodacom_u, 'vodacom_c': vodacom_c,
                    'orange_u': orange_u, 'orange_c': orange_c,
                    'pepele_u': pepele_u, 'pepele_c': pepele_c,
                    'cash_u': cash_u, 'cash_c': cash_c,
                }
            rows_imported += 1

        print(f'{guichetier.full_name}: {rows_imported} clotures imported')

        # Update per-guichetier virtual stocks and cash
        if last_data:
            for op_name, key_u, key_c in [
                ('Airtel Money', 'airtel_u', 'airtel_c'),
                ('M-Pesa', 'vodacom_u', 'vodacom_c'),
                ('Orange Money', 'orange_u', 'orange_c'),
                ('Pepete Mobile', 'pepele_u', 'pepele_c'),
            ]:
                op_type = OperationType.query.filter_by(name=op_name).first()
                if not op_type:
                    continue
                for cur, lk in [('USD', key_u), ('FC', key_c)]:
                    val = last_data[lk]
                    vs = VirtualStock.query.filter_by(agency_id=agency.id, operation_type_id=op_type.id, currency=cur, user_id=guichetier.id).first()
                    if vs:
                        vs.current_balance = val
                        vs.opening_balance = val
                    else:
                        vs = VirtualStock(agency_id=agency.id, operation_type_id=op_type.id, currency=cur, user_id=guichetier.id,
                                          opening_balance=val, current_balance=val)
                        db.session.add(vs)

            for cur, lk in [('USD', 'cash_u'), ('FC', 'cash_c')]:
                val = last_data[lk]
                cb = CashBalance.query.filter_by(agency_id=agency.id, currency=cur, user_id=guichetier.id).first()
                if cb:
                    cb.current_balance = val
                    cb.opening_balance = val
                else:
                    cb = CashBalance(agency_id=agency.id, currency=cur, user_id=guichetier.id,
                                     opening_balance=val, current_balance=val)
                    db.session.add(cb)

    # Also update agency-level stocks/cash from per-guichetier totals (sum of all guichetiers)
    for op_name in ['Airtel Money', 'M-Pesa', 'Orange Money', 'Pepete Mobile']:
        op_type = OperationType.query.filter_by(name=op_name).first()
        if not op_type:
            continue
        for cur in ['USD', 'FC']:
            total = db.session.query(db.func.sum(VirtualStock.current_balance)).filter_by(
                agency_id=agency.id, operation_type_id=op_type.id, currency=cur
            ).filter(VirtualStock.user_id.isnot(None)).scalar() or 0
            vs = VirtualStock.query.filter_by(agency_id=agency.id, operation_type_id=op_type.id, currency=cur, user_id=None).first()
            if vs:
                vs.current_balance = total
                vs.opening_balance = total
            else:
                vs = VirtualStock(agency_id=agency.id, operation_type_id=op_type.id, currency=cur, user_id=None,
                                  opening_balance=total, current_balance=total)
                db.session.add(vs)

    for cur in ['USD', 'FC']:
        total = db.session.query(db.func.sum(CashBalance.current_balance)).filter_by(
            agency_id=agency.id, currency=cur
        ).filter(CashBalance.user_id.isnot(None)).scalar() or 0
        cb = CashBalance.query.filter_by(agency_id=agency.id, currency=cur, user_id=None).first()
        if cb:
            cb.current_balance = total
            cb.opening_balance = total
        else:
            cb = CashBalance(agency_id=agency.id, currency=cur, user_id=None,
                             opening_balance=total, current_balance=total)
            db.session.add(cb)

    db.session.commit()
    print('Import complete!')
